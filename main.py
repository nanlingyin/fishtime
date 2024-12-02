# main2.py
import sys
import os
import json
import sqlite3
import ctypes
import win32gui
import win32con
import win32ui
from datetime import datetime
from ctypes import wintypes
from functools import lru_cache
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QLabel, QPushButton, QVBoxLayout,
    QWidget, QHBoxLayout, QSystemTrayIcon, QMenu, QAction, QColorDialog,
    QFileDialog, QMessageBox, QDialog, QTableWidget, QTableWidgetItem,
    QDialogButtonBox, QFormLayout, QFontDialog, QComboBox, QListWidget, QGroupBox, QInputDialog,QFileIconProvider
)
from PyQt5.QtCore import QTimer, Qt, QSize
from PyQt5.QtGui import QFont, QPixmap, QIcon, QColor, QPalette, QImage
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib import pyplot as plt
import psutil
import openpyxl
from fpdf import FPDF
os.environ['QT_QPA_PLATFORM_PLUGIN_PATH'] = r'C:\Users\Administrator\Desktop\fish_time\.venv\Lib\site-packages\PyQt5\Qt5\plugins\platforms'
# 环境变量必须保留！！！！！！！
from PyQt5.QtCore import QTimer, Qt, QSize
from PyQt5.QtGui import QFont, QPixmap, QIcon, QColor, QPalette, QImage
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib import pyplot as plt
from PyQt5.QtCore import QTimer, Qt, QSize, QFileInfo
from PyQt5.QtGui import QFont, QPixmap, QIcon, QColor, QPalette, QImage
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib import pyplot as plt

CONFIG_FILE = 'config.json'
DEFAULT_CONFIG = {
    "font_family": "Microsoft YaHei",
    "font_color": "#2E3440",
    "language": "zh_CN",
    "theme": "灰色",
    "Process Whitelist": []
}

LANGUAGES = {
    "zh_CN": {
        "fish": "总摸鱼时长: ",
        "current_app": "当前应用: {}",
        "no_usage": "今日无使用记录。",
        "export_success": "报告已保存至 {}",
        "export_error": "保存文件时出错: {}",
        "clear_confirm_title": "确认清除",
        "clear_confirm_text": "确定要清除所有记录吗？",
        "clear_success": "记录已清除。",
        "font_settings": "字体设置",
        "report": "生成报告",
        "clear": "清除记录",
        "font_setting_button": "字体设置",
        "tray_show": "显示",
        "tray_exit": "退出",
        "tray_minimized": "应用已最小化到系统托盘。",
        "information": "信息",
        "error": "错误",
        "application": "应用/进程",
        "duration_seconds": "使用时长（秒）",
        "duration_minutes": "使用时长（分钟）",
        "duration_hours": "使用时长（小时）",
        "start_tracking": "开始跟踪",
        "stop_tracking": "停止跟踪",
        "export_csv": "导出为 CSV",
        "export_pdf": "导出为 PDF",
        "dark_mode": "暗黑模式",
        "chart": "图表",
        "none": "无",
        "process_whitelist": "进程白名单",
        "add_process": "添加进程",
        "remove_selected": "移除选中"
    },
    "en_US": {
        "fish": "Slackoff Time：",
        "current_app": "Current Application: {}",
        "no_usage": "No usage records today.",
        "export_success": "Report saved to {}",
        "export_error": "Error saving file: {}",
        "clear_confirm_title": "Confirm Clear",
        "clear_confirm_text": "Are you sure you want to clear all records?",
        "clear_success": "Records cleared.",
        "font_settings": "Font Settings",
        "report": "Generate Report",
        "clear": "Clear Records",
        "font_setting_button": "Font Settings",
        "tray_show": "Show",
        "tray_exit": "Exit",
        "tray_minimized": "Application minimized to system tray.",
        "information": "Information",
        "error": "Error",
        "application": "Application/Process",
        "duration_seconds": "Usage Duration (seconds)",
        "duration_minutes": "Usage Duration (minutes)",
        "duration_hours": "Usage Duration (hours)",
        "start_tracking": "Start Tracking",
        "stop_tracking": "Stop Tracking",
        "export_csv": "Export as CSV",
        "export_pdf": "Export as PDF",
        "dark_mode": "Dark Mode",
        "chart": "Chart",
        "none": "none",
        "process_whitelist": "Process Whitelist",
        "add_process": "Add Process",
        "remove_selected": "Remove Selected"
    }
}

THEMES = {
    "灰色": """
        QMainWindow {{
            background-color: #A0AEC0;
            font-family: "{font_family}";
            color: {font_color};
        }}
        QPushButton {{
            border: none;
            border-radius: 8px;
            padding: 12px 24px;
            background-color: #5A67D8;
            color: white;
            font-size: 16px;
            font-family: "{font_family}";
        }}
        QPushButton:hover {{
            background-color: #434190;
        }}
        QPushButton:pressed {{
            background-color: #2B6CB0;
        }}
        QDialog {{
            background-color: #FFFFFF;
            font-family: "{font_family}";
        }}
        QLabel {{
            font-size: 18px;
            font-family: "{font_family}";
        }}
        QTableWidget {{
            font-family: "{font_family}";
            font-size: 14px;
            alternatingRowColors: true;
            background-color: #FFFFFF;
            gridline-color: #CBD5E0;
        }}
        QTableWidget::item:selected {{
            background-color: #A0AEC0;
            color: white;
        }}
    """,
    "白色": """
        QMainWindow {{
            background-color: #FFFFFF;
            font-family: "{font_family}";
            color: {font_color};
        }}
        QPushButton {{
            border: 1px solid #CCCCCC;
            border-radius: 8px;
            padding: 12px 24px;
            background-color: #F0F0F0;
            color: black;
            font-size: 16px;
            font-family: "{font_family}";
        }}
        QPushButton:hover {{
            background-color: #E0E0E0;
        }}
        QPushButton:pressed {{
            background-color: #D0D0D0;
        }}
        QDialog {{
            background-color: #FFFFFF;
            font-family: "{font_family}";
        }}
        QLabel {{
            font-size: 18px;
            font-family: "{font_family}";
        }}
        QTableWidget {{
            font-family: "{font_family}";
            font-size: 14px;
            alternatingRowColors: true;
            background-color: #FFFFFF;
            gridline-color: #CCCCCC;
        }}
        QTableWidget::item:selected {{
            background-color: #D0D0D0;
            color: black;
        }}
    """,
    "暗黑": """
        QMainWindow {{
            background-color: #2D3748;
            font-family: "{font_family}";
            color: {font_color};
        }}
        QPushButton {{
            border: none;
            border-radius: 8px;
            padding: 12px 24px;
            background-color: #4A5568;
            color: white;
            font-size: 16px;
            font-family: "{font_family}";
            font-weight: bold;
        }}
        QPushButton:hover {{
            background-color: #2D3748;
        }}
        QPushButton:pressed {{
            background-color: #1A202C;
        }}
        QDialog {{
            background-color: #1A202C;
            font-family: "{font_family}";
        }}
        QLabel {{
            font-size: 18px;
            font-family: "{font_family}";
            color: {font_color};
        }}
        QTableWidget {{
            font-family: "{font_family}";
            font-size: 14px;
            alternatingRowColors: true;
            background-color: #2D3748;
            gridline-color: #4A5568;
        }}
        QTableWidget::item:selected {{
            background-color: #4A5568;
            color: white;
        }}
    """
}

# 数据库初始化
def init_db():
    conn = sqlite3.connect('usage.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS usage (
            id INTEGER PRIMARY KEY,
            application TEXT,
            duration INTEGER,
            date TEXT
        )
    ''')
    conn.commit()
    conn.close()

# 加载配置
def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            try:
                config = json.load(f)
                # Ensure all default keys exist
                for key, value in DEFAULT_CONFIG.items():
                    if key not in config:
                        config[key] = value
                return config
            except json.JSONDecodeError:
                return DEFAULT_CONFIG.copy()
    else:
        config = DEFAULT_CONFIG.copy()
    for key, value in DEFAULT_CONFIG.items():
        if key not in config:
            config[key] = value
    return config

# 保存配置
def save_config(config):
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=4)

# 忽略列表，添加不需要监控的进程名称
IGNORE_LIST = [
    "Idle", "System", "Registry", "Memory Compression",
    "摸鱼时间记录.exe", "python.exe",
    "选择背景图片", "选择字体颜色", "选择字体", "编辑背景图片",
    "确认清除", "完成", "字体设置", "Figure"
]

APP_NAME = "摸鱼时间记录.exe"  # 确保这是应用的实际进程名称

# 数据库记录
def log_usage(app_name, duration):
    # 如果应用在忽略列表中，不记录
    if app_name in IGNORE_LIST or app_name == APP_NAME:
        return
    conn = sqlite3.connect('usage.db')
    cursor = conn.cursor()
    date_str = datetime.now().strftime("%Y-%m-%d")
    cursor.execute('''
        INSERT INTO usage (application, duration, date)
        VALUES (?, ?, ?)
    ''', (app_name, duration, date_str))
    conn.commit()
    conn.close()

# 进程图标缓存
icon_cache = {}

@lru_cache(maxsize=128)
def get_process_icon(app_name):
    """
    使用Win32 API获取进程的图标。如果无法获取，则返回默认图标。
    """
    if app_name in icon_cache:
        return icon_cache[app_name]

    try:
        for proc in psutil.process_iter(['name', 'exe']):
            if proc.info['name'] == app_name:
                exe_path = proc.info['exe']
                if exe_path and os.path.exists(exe_path):
                    icon_provider = QFileIconProvider()
                    qt_icon = icon_provider.icon(QFileInfo(exe_path))
                    pixmap = qt_icon.pixmap(64, 64)
                    icon_cache[app_name] = pixmap
                    return pixmap
    except Exception as e:
        print(f"获取 {app_name} 图标时出错: {e}")


    # 返回默认图标
    default_icon = QPixmap(64, 64)
    default_icon.fill(Qt.gray)
    icon_cache[app_name] = default_icon
    return default_icon

def hicon_to_qpixmap(hicon):
    """
    将 HICON 转换为 QPixmap
    """
    # 获取 icon info
    info = win32gui.GetIconInfo(hicon)
    bmp = info[2]
    width = win32gui.GetSystemMetrics(win32con.SM_CXICON)
    height = win32gui.GetSystemMetrics(win32con.SM_CYICON)

    # Create a device context
    hdc = win32gui.CreateCompatibleDC(0)
    hbmp = win32ui.CreateBitmap()
    hbmp.CreateCompatibleBitmap(win32ui.CreateDCFromHandle(hdc), width, height)
    hdc_obj = win32ui.CreateDCFromHandle(hdc)
    hdc_obj.SelectObject(hbmp)
    win32gui.DrawIconEx(hdc, 0, 0, hicon, width, height, 0, None, win32con.DI_NORMAL)

    # Get bitmap bits
    bmp_info = hbmp.GetInfo()
    bmp_str = hbmp.GetBitmapBits(True)

    # Create QImage from bitmap bits
    image = QImage(bmp_str, bmp_info['bmWidth'], bmp_info['bmHeight'], QImage.Format_ARGB32)
    pixmap = QPixmap.fromImage(image)

    # Clean up
    win32gui.DeleteObject(bmp)
    win32gui.DeleteObject(info[1])  # hbmColor
    win32gui.DeleteObject(info[0])  # hbmMask
    win32gui.DeleteDC(hdc)
    hdc_obj.DeleteDC()

    return pixmap

# 生成表格报告并导出为xlsx、csv和pdf
def generate_report(language):
    conn = sqlite3.connect('usage.db')
    cursor = conn.cursor()
    cursor.execute('''
        SELECT application, SUM(duration) FROM usage
        WHERE date=?
        GROUP BY application
        ORDER BY SUM(duration) DESC
    ''', (datetime.now().strftime("%Y-%m-%d"),))
    data = cursor.fetchall()
    conn.close()
    if not data:
        QMessageBox.information(None, LANGUAGES[language]["information"], LANGUAGES[language]["no_usage"])
        return

    # 创建表格对话框
    table_dialog = QDialog()
    table_dialog.setWindowTitle(LANGUAGES[language]["report"])
    table_dialog.setGeometry(200, 200, 900, 600)
    layout = QVBoxLayout()
    table_dialog.setLayout(layout)

    table = QTableWidget()
    table.setRowCount(len(data))
    table.setColumnCount(5)  # 应用名称，使用时长（秒），使用时长（分钟），使用时长（小时），图标
    table.setHorizontalHeaderLabels([
        LANGUAGES[language]["application"],
        LANGUAGES[language]["duration_seconds"],
        LANGUAGES[language]["duration_minutes"],
        LANGUAGES[language]["duration_hours"],
        "图标" if language == "zh_CN" else "Icon"
    ])
    table.setIconSize(QSize(64, 64))  # 设置图标尺寸
    table.setSelectionBehavior(QTableWidget.SelectRows)
    table.setEditTriggers(QTableWidget.NoEditTriggers)
    max_duration = max([duration for _, duration in data]) if data else 1

    for row, (app, duration) in enumerate(data):
        if app and app not in IGNORE_LIST:
            # 应用名称
            app_item = QTableWidgetItem(app)
            table.setItem(row, 0, app_item)

            # 使用时长（秒）
            table.setItem(row, 1, QTableWidgetItem(str(duration)))

            # 使用时长（分钟）
            minutes = duration / 60
            table.setItem(row, 2, QTableWidgetItem(f"{minutes:.2f}"))

            # 使用时长（小时）
            hours = duration / 3600
            table.setItem(row, 3, QTableWidgetItem(f"{hours:.2f}"))

            # 获取并设置应用图标
            icon_pixmap = get_process_icon(app)
            icon_label = QLabel()
            icon_label.setPixmap(icon_pixmap.scaled(48, 48, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            icon_label.setAlignment(Qt.AlignCenter)
            table.setCellWidget(row, 4, icon_label)

            # 设置单元格背景颜色根据使用时长
            intensity = min(int((duration / max_duration) * 255), 255)
            color = QColor( intensity,255 - intensity, 55)  # 从红到绿
            for col in range(4):
                table.item(row, col).setBackground(color)


    table.resizeColumnsToContents()
    table.horizontalHeader().setStretchLastSection(True)
    table.verticalHeader().setDefaultSectionSize(70)
    layout.addWidget(table)

    # 导出按钮布局
    export_layout = QHBoxLayout()
    export_layout.addStretch()

    # 导出为 XLSX
    export_xlsx_button = QPushButton("导出为 XLSX" if language == "zh_CN" else "Export as XLSX")
    export_xlsx_button.setFixedHeight(50)
    export_xlsx_button.setStyleSheet("""
        QPushButton {
            background-color: #5A67D8;
            color: white;
            border-radius: 8px;
            padding: 10px 20px;
            font-size: 18px;
            font-family: "Microsoft YaHei";
        }
        QPushButton:hover {
            background-color: #434190;
        }
    """)
    export_xlsx_button.clicked.connect(lambda: export_to_xlsx(data, language))
    export_layout.addWidget(export_xlsx_button)

    # 导出为 CSV
    export_csv_button = QPushButton(LANGUAGES[language]["export_csv"])
    export_csv_button.setFixedHeight(50)
    export_csv_button.setStyleSheet("""
        QPushButton {
            background-color: #38A169;
            color: white;
            border-radius: 8px;
            padding: 10px 20px;
            font-size: 18px;
            font-family: "Microsoft YaHei";
        }
        QPushButton:hover {
            background-color: #2F855A;
        }
    """)
    export_csv_button.clicked.connect(lambda: export_to_csv(data, language))
    export_layout.addWidget(export_csv_button)

    # 导出为 PDF
    export_pdf_button = QPushButton(LANGUAGES[language].get("export_pdf", "Export as PDF"))
    export_pdf_button.setFixedHeight(50)
    export_pdf_button.setStyleSheet("""
        QPushButton {
            background-color: #DD6B20;
            color: white;
            border-radius: 8px;
            padding: 10px 20px;
            font-size: 18px;
            font-family: "Microsoft YaHei";
        }
        QPushButton:hover {
            background-color: #C05621;
        }
    """)
    export_pdf_button.clicked.connect(lambda: export_to_pdf(data, language))
    export_layout.addWidget(export_pdf_button)

    layout.addLayout(export_layout)

    table_dialog.exec_()

def show_chart(app_name, language):
    conn = sqlite3.connect('usage.db')
    cursor = conn.cursor()
    cursor.execute('''
        SELECT date, SUM(duration) FROM usage
        WHERE application=?
        GROUP BY date
        ORDER BY date
    ''', (app_name,))
    data = cursor.fetchall()
    conn.close()

    if not data:
        QMessageBox.information(None, LANGUAGES[language]["information"], LANGUAGES[language]["no_usage"])
        return

    dates = [row[0] for row in data]
    durations = [row[1] for row in data]

    fig, ax = plt.subplots(figsize=(8, 6))
    ax.plot(dates, durations, marker='o', linestyle='-')
    ax.set_title(f"{app_name} 使用时长" if language == "zh_CN" else f"Usage Duration of {app_name}")
    ax.set_xlabel("日期" if language == "zh_CN" else "Date")
    ax.set_ylabel("秒数" if language == "zh_CN" else "Seconds")
    plt.xticks(rotation=45)
    plt.tight_layout()

    chart_dialog = QDialog()
    chart_dialog.setWindowTitle(f"{app_name} {LANGUAGES[language]['report']}")
    chart_dialog.setGeometry(250, 250, 800, 600)
    layout = QVBoxLayout()
    chart_dialog.setLayout(layout)

    canvas = FigureCanvas(fig)
    layout.addWidget(canvas)

    chart_dialog.exec_()

# 导出数据为xlsx
def export_to_xlsx(data, language):
    options = QFileDialog.Options()
    file_path, _ = QFileDialog.getSaveFileName(None, "Save as XLSX" if language == "en_US" else "保存为 XLSX", "", "Excel Files (*.xlsx)", options=options)
    if file_path:
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Usage Report"
            ws.append([
                LANGUAGES[language]["application"],
                LANGUAGES[language]["duration_seconds"],
                LANGUAGES[language]["duration_minutes"],
                LANGUAGES[language]["duration_hours"]
            ])
            for row in data:
                if row[0] and row[0] not in IGNORE_LIST:
                    duration_seconds = row[1]
                    duration_minutes = duration_seconds / 60
                    duration_hours = duration_seconds / 3600
                    ws.append([row[0], duration_seconds, f"{duration_minutes:.2f}", f"{duration_hours:.2f}"])
            wb.save(file_path)
            QMessageBox.information(None, LANGUAGES[language]["information"], LANGUAGES[language]["export_success"].format(file_path))
        except Exception as e:
            QMessageBox.critical(None, LANGUAGES[language]["error"], LANGUAGES[language]["export_error"].format(e))

# 导出数据为csv
def export_to_csv(data, language):
    options = QFileDialog.Options()
    file_path, _ = QFileDialog.getSaveFileName(None, "Save as CSV" if language == "en_US" else "保存为 CSV", "", "CSV Files (*.csv)", options=options)
    if file_path:
        try:
            import csv
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow([
                    LANGUAGES[language]["application"],
                    LANGUAGES[language]["duration_seconds"],
                    LANGUAGES[language]["duration_minutes"],
                    LANGUAGES[language]["duration_hours"]
                ])
                for row in data:
                    if row[0] and row[0] not in IGNORE_LIST:
                        duration_seconds = row[1]
                        duration_minutes = duration_seconds / 60
                        duration_hours = duration_seconds / 3600
                        writer.writerow([row[0], duration_seconds, f"{duration_minutes:.2f}", f"{duration_hours:.2f}"])
            QMessageBox.information(None, LANGUAGES[language]["information"], LANGUAGES[language]["export_success"].format(file_path))
        except Exception as e:
            QMessageBox.critical(None, LANGUAGES[language]["error"], LANGUAGES[language]["export_error"].format(e))

# 导出数据为pdf
def export_to_pdf(data, language):
    options = QFileDialog.Options()
    file_path, _ = QFileDialog.getSaveFileName(None, "Save as PDF" if language == "en_US" else "保存为 PDF", "", "PDF Files (*.pdf)", options=options)
    if file_path:
        try:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(0, 10, LANGUAGES[language]["report"], ln=True, align='C')
            pdf.set_font("Arial", 'B', 12)
            headers = [
                LANGUAGES[language]["application"],
                LANGUAGES[language]["duration_seconds"],
                LANGUAGES[language]["duration_minutes"],
                LANGUAGES[language]["duration_hours"]
            ]
            for header in headers:
                pdf.cell(50, 10, header, border=1, align='C')
            pdf.ln()
            pdf.set_font("Arial", '', 12)
            for row in data:
                if row[0] and row[0] not in IGNORE_LIST:
                    duration_seconds = row[1]
                    duration_minutes = duration_seconds / 60
                    duration_hours = duration_seconds / 3600
                    pdf.cell(50, 10, row[0], border=1, align='C')
                    pdf.cell(50, 10, str(duration_seconds), border=1, align='C')
                    pdf.cell(50, 10, f"{duration_minutes:.2f}", border=1, align='C')
                    pdf.cell(50, 10, f"{duration_hours:.2f}", border=1, align='C')
                    pdf.ln()
            pdf.output(file_path)
            QMessageBox.information(None, LANGUAGES[language]["information"], LANGUAGES[language]["export_success"].format(file_path))
        except Exception as e:
            QMessageBox.critical(None, LANGUAGES[language]["error"], LANGUAGES[language]["export_error"].format(e))

# 清除记录
def clear_records(language):
    if os.path.exists('usage.db'):
        try:
            conn = sqlite3.connect('usage.db')
            cursor = conn.cursor()
            cursor.execute('DELETE FROM usage')
            conn.commit()
            conn.close()
        except Exception as e:
            QMessageBox.critical(None, LANGUAGES[language]["error"], LANGUAGES[language]["export_error"].format(e))
    init_db()
    QMessageBox.information(None, LANGUAGES[language]["information"], LANGUAGES[language]["clear_success"])

# 使用 ctypes 获取活动窗口的窗口标题和句柄
user32 = ctypes.WinDLL('user32', use_last_error=True)

def get_active_window():
    hwnd = user32.GetForegroundWindow()
    length = user32.GetWindowTextLengthW(hwnd)
    buffer = ctypes.create_unicode_buffer(length + 1)
    user32.GetWindowTextW(hwnd, buffer, length + 1)
    window_title = buffer.value
    if window_title:
        pid = wintypes.DWORD()
        user32.GetWindowThreadProcessId(hwnd, ctypes.byref(pid))
        try:
            process = psutil.Process(pid.value)
            return process.name()
        except Exception:
            return None
    else:
        return None

# 字体选择对话框
class FontSettingsDialog(QDialog):
    def __init__(self, current_font, current_color, parent=None):
        super().__init__(parent)
        self.setWindowTitle("字体设置" if parent.current_language == "zh_CN" else "Font Settings")
        self.setFixedSize(500, 300)
        self.selected_font = current_font
        self.selected_color = QColor(current_color)

        self.layout = QFormLayout()
        self.setLayout(self.layout)

        # 字体选择
        self.font_button = QPushButton("选择字体" if parent.current_language == "zh_CN" else "Choose Font")
        self.font_button.clicked.connect(self.choose_font)
        self.layout.addRow("字体:" if parent.current_language == "zh_CN" else "Font:", self.font_button)

        # 字体颜色选择
        self.color_button = QPushButton("选择颜色" if parent.current_language == "zh_CN" else "Choose Color")
        self.color_button.clicked.connect(self.choose_color)
        self.layout.addRow("字体颜色:" if parent.current_language == "zh_CN" else "Font Color:", self.color_button)

        self.button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        self.layout.addWidget(self.button_box)

    def choose_font(self):
        dialog = QFontDialog(self)
        dialog.setOption(QFontDialog.DontUseNativeDialog, True)
        dialog.setCurrentFont(self.selected_font)
        dialog.setWindowTitle("选择字体" if self.parent().current_language == "zh_CN" else "Choose Font")
        if dialog.exec_() == QDialog.Accepted:
            self.selected_font = dialog.selectedFont()

    def choose_color(self):
        dialog_title = "选择字体颜色" if self.parent().current_language == "zh_CN" else "Choose Font Color"
        color = QColorDialog.getColor(self.selected_color, self, dialog_title)
        if color.isValid():
            self.selected_color = color

    def get_settings(self):
        return self.selected_font, self.selected_color.name()

# 自定义标题栏（带图标）
class TitleBar(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.setAutoFillBackground(True)
        self.setBackgroundRole(QPalette.Window)

        self.layout = QHBoxLayout()
        self.layout.setContentsMargins(10, 0, 10, 0)
        self.setLayout(self.layout)
        self.setFixedHeight(50)

        # 设置应用图标
        icon_path = r'C:\Users\Administrator\Desktop\摸鱼时间记录\icon\text.png'
        if os.path.exists(icon_path):
            self.icon_label = QLabel()
            pixmap = QPixmap(icon_path).scaled(30, 30, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.icon_label.setPixmap(pixmap)
            self.layout.addWidget(self.icon_label)
        else:
            self.icon_label = QLabel()
            self.layout.addWidget(self.icon_label)

        # 设置标题
        self.title = QLabel("摸鱼时间记录" if self.parent.current_language == "zh_CN" else "Time Tracker")
        self.title.setFont(QFont(self.parent.config.get("font_family", "Microsoft YaHei"), 16))
        self.layout.addWidget(self.title)

        self.layout.addStretch()

        # 窗口控制按钮
        button_size = 35

        self.min_button = QPushButton("")
        self.min_button.setFixedSize(button_size, button_size)
        self.min_button.setStyleSheet(circle_button_style("#FFBD44"))
        self.min_button.setToolTip("最小化" if self.parent.current_language == "zh_CN" else "Minimize")
        self.min_button.clicked.connect(self.parent.showMinimized)
        self.min_button.setCursor(Qt.PointingHandCursor)
        self.layout.addWidget(self.min_button)

        self.max_button = QPushButton("")
        self.max_button.setFixedSize(button_size, button_size)
        self.max_button.setStyleSheet(circle_button_style("#5A67D8"))
        self.max_button.setToolTip("最大化" if self.parent.current_language == "zh_CN" else "Maximize")
        self.max_button.clicked.connect(self.toggle_maximize)
        self.max_button.setCursor(Qt.PointingHandCursor)
        self.layout.addWidget(self.max_button)

        self.close_button = QPushButton("")
        self.close_button.setFixedSize(button_size, button_size)
        self.close_button.setStyleSheet(circle_button_style("#E53E3E"))
        self.close_button.setToolTip("关闭" if self.parent.current_language == "zh_CN" else "Close")
        self.close_button.clicked.connect(self.parent.close)
        self.close_button.setCursor(Qt.PointingHandCursor)
        self.layout.addWidget(self.close_button)

        self.start = None

    def toggle_maximize(self):
        if self.parent.isMaximized():
            self.parent.showNormal()
            self.max_button.setText("")
            self.max_button.setStyleSheet(circle_button_style("#5A67D8"))
            self.max_button.setToolTip("最大化" if self.parent.current_language == "zh_CN" else "Maximize")
        else:
            self.parent.showMaximized()
            self.max_button.setText("")
            self.max_button.setStyleSheet(circle_button_style("#63B3ED"))
            self.max_button.setToolTip("还原" if self.parent.current_language == "zh_CN" else "Restore")

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.start = event.globalPos()
            self.click_position = event.pos()

    def mouseMoveEvent(self, event):
        if self.start:
            delta = event.globalPos() - self.start
            self.parent.move(self.parent.pos() + delta)
            self.start = event.globalPos()

    def mouseReleaseEvent(self, event):
        self.start = None

def circle_button_style(color):
    return f"""
        QPushButton {{
            background-color: {color};
            border: none;
            border-radius: 17.5px;
            margin: 5px;
        }}
        QPushButton:hover {{
            border: 2px solid #FFFFFF;
            opacity: 0.9;
        }}
        QPushButton:pressed {{
            background-color: {color};
        }}
    """

# 主窗口
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.config = load_config()
        self.current_language = self.config.get("language", "zh_CN")
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setStyleSheet(THEMES[self.config.get("theme", "灰色")].format(
            font_family=self.config.get("font_family", "Microsoft YaHei"),
            font_color=self.config.get("font_color", "#2E3440")
        ))

        self.title_bar = TitleBar(self)
        self.setCentralWidget(QWidget())
        self.main_layout = QVBoxLayout()
        self.centralWidget().setLayout(self.main_layout)
        self.main_layout.addWidget(self.title_bar)

        # 添加内容区域
        self.content = QWidget()
        self.content_layout = QVBoxLayout()
        self.content.setLayout(self.content_layout)
        self.main_layout.addWidget(self.content)

        # 总摸鱼时长标签
        self.total_time_label = QLabel()
        self.total_time_label.setAlignment(Qt.AlignCenter)
        self.total_time_label.setFont(QFont(self.config.get("font_family", "Microsoft YaHei"), 14))
        self.content_layout.addWidget(self.total_time_label)

        # 当前应用名称和图标
        self.current_app_label = QLabel(LANGUAGES[self.current_language]["current_app"].format("无"))
        self.current_app_label.setAlignment(Qt.AlignCenter)
        self.current_app_label.setFont(QFont(self.config.get("font_family", "Microsoft YaHei"), 14))
        self.content_layout.addWidget(self.current_app_label)

        self.current_icon_label = QLabel()
        default_pixmap = QPixmap(64, 64)
        default_pixmap.fill(Qt.gray)
        self.current_icon_label.setPixmap(default_pixmap)
        self.current_icon_label.setAlignment(Qt.AlignCenter)
        self.content_layout.addWidget(self.current_icon_label)

        # 跟踪控制按钮
        self.toggle_tracking_button = QPushButton(LANGUAGES[self.current_language]["start_tracking"])
        self.toggle_tracking_button.clicked.connect(self.toggle_tracking)
        self.content_layout.addWidget(self.toggle_tracking_button)

        # 生成报告按钮
        self.report_button = QPushButton(LANGUAGES[self.current_language]["report"])
        self.report_button.clicked.connect(lambda: generate_report(self.current_language))
        self.content_layout.addWidget(self.report_button)

        # 清除记录按钮
        self.clear_button = QPushButton(LANGUAGES[self.current_language]["clear"])
        self.clear_button.clicked.connect(self.clear_records_confirmation)
        self.content_layout.addWidget(self.clear_button)

        # 字体设置按钮
        self.font_settings_button = QPushButton(LANGUAGES[self.current_language]["font_settings"])
        self.font_settings_button.clicked.connect(self.open_font_settings)
        self.content_layout.addWidget(self.font_settings_button)

        # 语言选择
        self.language_combo = QComboBox()
        self.language_combo.addItems(["中文", "English"])
        self.language_combo.setCurrentIndex(0 if self.current_language == "zh_CN" else 1)
        self.language_combo.currentIndexChanged.connect(self.change_language)
        self.content_layout.addWidget(self.language_combo)

        # 主题选择
        self.theme_combo = QComboBox()
        self.theme_combo.addItems(["暗黑", "白色", "灰色"])
        self.theme_combo.setCurrentText(self.config.get("theme", "灰色"))
        self.theme_combo.currentIndexChanged.connect(self.change_theme)
        self.content_layout.addWidget(self.theme_combo)

        # 进程白名单管理
        self.whitelist_group = QGroupBox(LANGUAGES[self.current_language]["process_whitelist"])
        self.whitelist_layout = QVBoxLayout()
        self.whitelist_group.setLayout(self.whitelist_layout)
        self.content_layout.addWidget(self.whitelist_group)

        self.whitelist_list = QListWidget()
        self.whitelist_layout.addWidget(self.whitelist_list)

        self.whitelist_buttons_layout = QHBoxLayout()

        # 添加进程按钮
        self.add_process_button = QPushButton(LANGUAGES[self.current_language]["add_process"])
        self.add_process_button.clicked.connect(self.add_to_whitelist)
        self.whitelist_buttons_layout.addWidget(self.add_process_button)

        # 移除选中按钮
        self.remove_process_button = QPushButton(LANGUAGES[self.current_language]["remove_selected"])
        self.remove_process_button.clicked.connect(self.remove_from_whitelist)
        self.whitelist_buttons_layout.addWidget(self.remove_process_button)

        self.whitelist_layout.addLayout(self.whitelist_buttons_layout)

        # 系统托盘
        self.create_tray_icon()

        # 跟踪定时器
        self.tracking_timer = QTimer()
        self.tracking_timer.timeout.connect(self.track_usage)
        self.tracking = False
        self.duration = 0

        # 更新当前应用显示
        self.update_current_app()
        # 加载白名单
        self.load_whitelist()
        # 定时更新当前应用
        self.app_update_timer = QTimer()
        self.app_update_timer.timeout.connect(self.update_current_app)
        self.app_update_timer.start(1000)  # 每秒更新一次
        # 初始化计时器
        self.init_total_time_timer()

        # 更新总摸鱼时长
        self.update_total_time()

    def load_whitelist(self):
        whitelist = self.config.get("Process Whitelist", [])
        for process in whitelist:
            self.whitelist_list.addItem(process)
            IGNORE_LIST.append(process)

    def add_to_whitelist(self):
        text, ok = QInputDialog.getText(
            self,
            LANGUAGES[self.current_language]["add_process"],
            LANGUAGES[self.current_language]["add_process"]
        )
        if ok and text:
            if text not in IGNORE_LIST:
                self.whitelist_list.addItem(text)
                IGNORE_LIST.append(text)
                self.config["Process Whitelist"].append(text)
                save_config(self.config)
            else:
                QMessageBox.warning(
                    self,
                    LANGUAGES[self.current_language]["error"],
                    f"{text} {LANGUAGES[self.current_language].get('already_in_whitelist', '已经在白名单中。')}"
                )

    def remove_from_whitelist(self):
        selected_items = self.whitelist_list.selectedItems()
        if not selected_items:
            return
        for item in selected_items:
            process = item.text()
            if process in IGNORE_LIST:
                IGNORE_LIST.remove(process)
            if process in self.config["Process Whitelist"]:
                self.config["Process Whitelist"].remove(process)
            self.whitelist_list.takeItem(self.whitelist_list.row(item))
        save_config(self.config)


    def init_total_time_timer(self):
        self.total_time_timer = QTimer()
        self.total_time_timer.timeout.connect(self.update_total_time)
        self.total_time_timer.start(1000)  # 每秒更新一次

    def update_total_time(self):
        total_seconds = self.get_total_usage_time()
        formatted_time = self.format_duration(total_seconds)
        self.total_time_label.setText(LANGUAGES[self.current_language]["fish"]+formatted_time)

    def get_total_usage_time(self):
        conn = sqlite3.connect('usage.db')
        cursor = conn.cursor()
        cursor.execute('SELECT SUM(duration) FROM usage')
        result = cursor.fetchone()
        conn.close()
        return result[0] if result[0] else 0

    def format_duration(self, seconds):
        intervals = (
            ('年'if self.current_language == "zh_CN" else "Y", 31536000),  # 365*24*60*60
            ('月'if self.current_language == "zh_CN" else "Mon", 8592000),   # 30*24*60*60
            ('天'if self.current_language == "zh_CN" else "D", 2592000),      # 24*60*60
            ('小时'if self.current_language == "zh_CN" else "H", 1280000),     # 60*60
            ('分钟'if self.current_language == "zh_CN" else "Min", 86400),
            ('秒'if self.current_language == "zh_CN" else "S", 1),
        )
        for name, count in intervals:
            value = seconds // count
            if value >= 1:
                return f"{int(value)}{name}"
        return "0秒"

    def apply_theme(self):
        self.setStyleSheet(THEMES[self.config.get("theme", "灰色")].format(
            font_family=self.config.get("font_family", "Microsoft YaHei"),
            font_color=self.config.get("font_color", "#2E3440")
        ))
        self.title_bar.title.setFont(QFont(self.config.get("font_family", "Microsoft YaHei"), 16))
        self.update_ui_language()

    def create_tray_icon(self):
        icon_path = r'C:\Users\Administrator\Desktop\摸鱼时间记录\icon\text.png'
        if os.path.exists(icon_path):
            tray_icon = QIcon(icon_path)
        else:
            tray_icon = QIcon()

        self.tray = QSystemTrayIcon()
        self.tray.setIcon(tray_icon)
        self.tray.setVisible(True)

        menu = QMenu()

        show_action = QAction(LANGUAGES[self.current_language]["tray_show"])
        show_action.triggered.connect(self.show_normal)
        menu.addAction(show_action)

        exit_action = QAction(LANGUAGES[self.current_language]["tray_exit"])
        exit_action.triggered.connect(self.exit_app)
        menu.addAction(exit_action)

        self.tray.setContextMenu(menu)
        self.tray.activated.connect(self.tray_icon_clicked)

    def track_usage(self):
        app_name = get_active_window()
        if app_name:
            log_usage(app_name, 1)  # 每秒记录一次
        self.duration += 1

    def closeEvent(self, event):
        event.ignore()
        self.hide()
        self.tray.showMessage(
            LANGUAGES[self.current_language]["information"],
            LANGUAGES[self.current_language]["tray_minimized"],
            QSystemTrayIcon.Information,
            2000
        )

    def tray_icon_clicked(self, reason):
        if reason == QSystemTrayIcon.Trigger:
            self.show_normal()

    def show_normal(self):
        self.show()
        self.tray.hide()

    def exit_app(self):
        QApplication.quit()

    def clear_records_confirmation(self):
        reply = QMessageBox.question(self, LANGUAGES[self.current_language]["clear_confirm_title"],
                                     LANGUAGES[self.current_language]["clear_confirm_text"],
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            clear_records(self.current_language)

    def open_font_settings(self):
        dialog = FontSettingsDialog(QFont(self.config.get("font_family", "Microsoft YaHei"), 12),
                                    self.config.get("font_color", "#2E3440"),
                                    self)
        if dialog.exec_() == QDialog.Accepted:
            selected_font, selected_color = dialog.get_settings()
            self.config["font_family"] = selected_font.family()
            self.config["font_color"] = selected_color
            save_config(self.config)
            self.apply_theme()

    def change_language(self, index):
        self.current_language = "zh_CN" if index == 0 else "en_US"
        self.config["language"] = self.current_language
        save_config(self.config)
        self.apply_theme()
        self.update_whitelist_ui_texts()
        self.apply_theme()
    


    def change_theme(self, index):
        themes = ["暗黑", "白色", "灰色"]
        selected_theme = themes[index]
        self.config["theme"] = selected_theme
        save_config(self.config)
        self.apply_theme()

    def update_ui_language(self):
        self.toggle_tracking_button.setText(
            LANGUAGES[self.current_language]["stop_tracking"] if self.tracking else LANGUAGES[self.current_language]["start_tracking"]
        )
        self.report_button.setText(LANGUAGES[self.current_language]["report"])
        self.clear_button.setText(LANGUAGES[self.current_language]["clear"])
        self.font_settings_button.setText(LANGUAGES[self.current_language]["font_settings"])
        self.language_combo.setItemText(0, "中文" if self.current_language == "zh_CN" else "Chinese")
        self.language_combo.setItemText(1, "English" if self.current_language == "en_US" else "英语")
        self.theme_combo.setItemText(0, "暗黑" if self.current_language == "zh_CN" else "Dark")
        self.theme_combo.setItemText(1, "白色" if self.current_language == "zh_CN" else "White")
        self.theme_combo.setItemText(2, "灰色" if self.current_language == "zh_CN" else "Gray")
        self.title_bar.title.setText("摸鱼时间记录" if self.current_language == "zh_CN" else "Time Tracker")
        self.title_bar.min_button.setToolTip("最小化" if self.current_language == "zh_CN" else "Minimize")
        self.title_bar.max_button.setToolTip("最大化" if self.current_language == "zh_CN" else "Maximize")
        self.title_bar.close_button.setToolTip("关闭" if self.current_language == "zh_CN" else "Close")
        self.current_app_label.setText(
            LANGUAGES[self.current_language]["current_app"].format("无" if self.current_language == "zh_CN" else "None")
        )
    def update_whitelist_ui_texts(self):
        self.whitelist_group.setTitle(LANGUAGES[self.current_language]["process_whitelist"])
        self.add_process_button.setText(LANGUAGES[self.current_language]["add_process"])
        self.remove_process_button.setText(LANGUAGES[self.current_language]["remove_selected"])

    def toggle_tracking(self):
        if self.tracking:
            self.tracking_timer.stop()
            self.tracking = False
            self.toggle_tracking_button.setText(LANGUAGES[self.current_language]["start_tracking"])
        else:
            self.tracking_timer.start(1000)
            self.tracking = True
            self.toggle_tracking_button.setText(LANGUAGES[self.current_language]["stop_tracking"])
        self.update_ui_language()

    def update_current_app(self):
        app_name = get_active_window()
        if app_name and app_name not in IGNORE_LIST:
            self.current_app_label.setText(LANGUAGES[self.current_language]["current_app"].format(app_name))
            icon_pixmap = get_process_icon(app_name)
            self.current_icon_label.setPixmap(icon_pixmap.scaled(64, 64, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        else:
            self.current_app_label.setText(LANGUAGES[self.current_language]["current_app"].format("无"if self.current_language == "zh_CN" else "None"))
            default_pixmap = QPixmap(64, 64)
            default_pixmap.fill(Qt.gray)
            self.current_icon_label.setPixmap(default_pixmap)

class BITMAPINFOHEADER(ctypes.Structure):
    _fields_ = [
        ('biSize', wintypes.DWORD),
        ('biWidth', wintypes.LONG),
        ('biHeight', wintypes.LONG),
        ('biPlanes', wintypes.WORD),
        ('biBitCount', wintypes.WORD),
        ('biCompression', wintypes.DWORD),
        ('biSizeImage', wintypes.DWORD),
        ('biXPelsPerMeter', wintypes.LONG),
        ('biYPelsPerMeter', wintypes.LONG),
        ('biClrUsed', wintypes.DWORD),
        ('biClrImportant', wintypes.DWORD),
    ]

class BITMAPINFO(ctypes.Structure):
    _fields_ = [
        ('bmiHeader', BITMAPINFOHEADER),
        ('bmiColors', wintypes.DWORD * 3),
    ]

# 初始化并运行应用程序
if __name__ == '__main__':
    try:
        import psutil
        import openpyxl
        from fpdf import FPDF
        import win32gui
        import win32con
        import win32ui
    except ImportError:
        print("缺少必要的依赖项，请运行以下命令安装：")
        print("pip install psutil PyQt5 matplotlib openpyxl fpdf pywin32")
        sys.exit(1)

    init_db()
    config = load_config()
    save_config(config)
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    window = MainWindow()
    window.show()
    sys.exit(app.exec_())