# utils.py
from datetime import datetime
import csv
import openpyxl
from fpdf import FPDF
from PyQt5.QtWidgets import QMessageBox, QFileDialog
from constants import LANGUAGES

def format_duration(seconds, language):
    """
    格式化时间，将秒数转换为更易读的格式
    """
    intervals = (
        ('年' if language == "zh_CN" else "Y", 31536000),  # 365*24*60*60
        ('月' if language == "zh_CN" else "Mon", 8592000),   # 30*24*60*60
        ('天' if language == "zh_CN" else "D", 2592000),      # 24*60*60
        ('小时' if language == "zh_CN" else "H", 1280000),     # 60*60
        ('分钟' if language == "zh_CN" else "Min", 86400),
        ('秒' if language == "zh_CN" else "S", 1),
    )
    for name, count in intervals:
        value = seconds // count
        if value >= 1:
            return f"{int(value)}{name}"
    return "0秒" if language == "zh_CN" else "0S"

def circle_button_style(color):
    """
    创建圆形按钮样式
    """
    return f"""
        QPushButton {{
            background-color: {color};
            border: none;
            border-radius: 17.5px;
            margin: 5px;
        }}
        QPushButton:hover {{
            border: 2px solid #FFFFFF;
            opacity: 0.9;
        }}
        QPushButton:pressed {{
            background-color: {color};
        }}
    """

def export_to_xlsx(data, language):
    """导出数据为Excel格式"""
    options = QFileDialog.Options()
    file_path, _ = QFileDialog.getSaveFileName(None, "Save as XLSX" if language == "en_US" else "保存为 XLSX", "", "Excel Files (*.xlsx)", options=options)
    if file_path:
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Usage Report"
            ws.append([
                LANGUAGES[language]["application"],
                LANGUAGES[language]["duration_seconds"],
                LANGUAGES[language]["duration_minutes"],
                LANGUAGES[language]["duration_hours"]
            ])
            for row in data:
                if row[0]:
                    duration_seconds = row[1]
                    duration_minutes = duration_seconds / 60
                    duration_hours = duration_seconds / 3600
                    ws.append([row[0], duration_seconds, f"{duration_minutes:.2f}", f"{duration_hours:.2f}"])
            wb.save(file_path)
            QMessageBox.information(None, LANGUAGES[language]["information"], LANGUAGES[language]["export_success"].format(file_path))
            return True
        except Exception as e:
            QMessageBox.critical(None, LANGUAGES[language]["error"], LANGUAGES[language]["export_error"].format(e))
            return False

def export_to_csv(data, language):
    """导出数据为CSV格式"""
    options = QFileDialog.Options()
    file_path, _ = QFileDialog.getSaveFileName(None, "Save as CSV" if language == "en_US" else "保存为 CSV", "", "CSV Files (*.csv)", options=options)
    if file_path:
        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow([
                    LANGUAGES[language]["application"],
                    LANGUAGES[language]["duration_seconds"],
                    LANGUAGES[language]["duration_minutes"],
                    LANGUAGES[language]["duration_hours"]
                ])
                for row in data:
                    if row[0]:
                        duration_seconds = row[1]
                        duration_minutes = duration_seconds / 60
                        duration_hours = duration_seconds / 3600
                        writer.writerow([row[0], duration_seconds, f"{duration_minutes:.2f}", f"{duration_hours:.2f}"])
            QMessageBox.information(None, LANGUAGES[language]["information"], LANGUAGES[language]["export_success"].format(file_path))
            return True
        except Exception as e:
            QMessageBox.critical(None, LANGUAGES[language]["error"], LANGUAGES[language]["export_error"].format(e))
            return False

def export_to_pdf(data, language):
    """导出数据为PDF格式"""
    options = QFileDialog.Options()
    file_path, _ = QFileDialog.getSaveFileName(None, "Save as PDF" if language == "en_US" else "保存为 PDF", "", "PDF Files (*.pdf)", options=options)
    if file_path:
        try:
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(0, 10, LANGUAGES[language]["report"], ln=True, align='C')
            pdf.set_font("Arial", 'B', 12)
            headers = [
                LANGUAGES[language]["application"],
                LANGUAGES[language]["duration_seconds"],
                LANGUAGES[language]["duration_minutes"],
                LANGUAGES[language]["duration_hours"]
            ]
            for header in headers:
                pdf.cell(50, 10, header, border=1, align='C')
            pdf.ln()
            pdf.set_font("Arial", '', 12)
            for row in data:
                if row[0]:
                    duration_seconds = row[1]
                    duration_minutes = duration_seconds / 60
                    duration_hours = duration_seconds / 3600
                    pdf.cell(50, 10, row[0], border=1, align='C')
                    pdf.cell(50, 10, str(duration_seconds), border=1, align='C')
                    pdf.cell(50, 10, f"{duration_minutes:.2f}", border=1, align='C')
                    pdf.cell(50, 10, f"{duration_hours:.2f}", border=1, align='C')
                    pdf.ln()
            pdf.output(file_path)
            QMessageBox.information(None, LANGUAGES[language]["information"], LANGUAGES[language]["export_success"].format(file_path))
            return True
        except Exception as e:
            QMessageBox.critical(None, LANGUAGES[language]["error"], LANGUAGES[language]["export_error"].format(e))
            return False